import time
import webbrowser as web
import pyautogui as pg

import openpyxl
from tkinter import ttk
from tkinter import *
import tkinter as tk
from tkinter import filedialog


class Product:

    data=[]

    def __init__(self, win):
        self.wind = win
        self.wind.title('Products Applications')

        # Funtion Add File
        def abrir_archivo():
            archivo = filedialog.askopenfilename(title="abrir", initialdir="c:/",
                                                 filetypes=[("Documentos Excel", "*.xlsx*")])
            print(archivo)
            imprimir_datos(archivo)

        # Create a Frame
        frame = LabelFrame(self.wind, text='Cargar Datos')
        frame.grid(row=0, column=0, columnspan=3, pady=20)

        # Button Add file in Frame
        ttk.Button(frame, text="Abrir Archivo", command=abrir_archivo).grid(row=0, columnspan=1, sticky=W + E)

        # label Previsualizacion
        frame2 = Label(self.wind, text='Previsualizacion de mensaje')
        frame2.grid(row=1, column=0, columnspan=3)

        # Cuadro Previsualizacion
        cuador_previsualizacion = Frame()
        cuador_previsualizacion.grid(row=2, column=0, columnspan=3)
        cuador_previsualizacion.config(width="500", height="50", bg="#E9E2E1")

        mess = tk.StringVar()
        mess.set('...')
        frame2 = Label(self.wind, textvariable=mess)
        frame2.grid(row=2, column=0, columnspan=3)
        frame2.config(bg="#E9E2E1")

        # Funtion Save Datos
        def cargar_datos(archivo):
            excel_dataframe = openpyxl.load_workbook(archivo)
            dataframe = excel_dataframe.active
            data = []

            for row in range(1, dataframe.max_row):
                _row = []

                for col in dataframe.iter_cols(1, dataframe.max_column):
                    _row.append(col[row].value)

                data.append(_row)
            print(data)
            print(len(data))
            self.data=data
            return data

        # Table
        self.tree = ttk.Treeview(height=5)
        self.tree["columns"] = ("one", "two", "three", "four")
        self.tree.grid(row=3, column=0, columnspan=2, pady=15)
        self.tree.heading('#0', text='Nombre', anchor=CENTER)
        self.tree.heading('#1', text='Propiedad', anchor=CENTER)
        self.tree.heading('#2', text='Direccion', anchor=CENTER)
        self.tree.heading('#3', text='Costo', anchor=CENTER)
        self.tree.heading('#4', text='Celular', anchor=CENTER)

        def imprimir_datos(datos):
            registros = cargar_datos(datos)

            # clean table
            records = self.tree.get_children()
            for element in records:
                self.tree.delete(element)
            for row in registros:
                self.tree.insert("", 0, text=row[0], values=[row[1], row[2], row[3], row[4]])
                mess.set('Hola ' + str(row[0]) + ', somos Bienes Raices Pizarro. Dado al interes por la \n' + str(row[1]) +
                         ' que se encuentra en' + str(row[0]) + 'tiene un valor de ' + str(row[3]))


        def enviar_mens():
            for i in self.data:
                mensaje = ('Hola ' + str(i[0]) + ', somos "Bienes Raices en Sueños". Dado al interes por la \n'
                           + str(i[1]) +' que se encuentra en ' + str(i[2]) + ' tiene un valor de ' + str(i[3]))
                print(mensaje)
                web.open("https://web.whatsapp.com/send?phone=" + str(i[4]) + "&text=" + mensaje)

                #tiempo de espera
                time.sleep(8)
                pg.click(1230,964)
                time.sleep(2)
                pg.press('enter')

                time.sleep(1.5)
                pg.hotkey('ctrl','w')
                time.sleep(1)



        ttk.Button(win, text="Enviar mensajes", command=enviar_mens).grid(row=4, columnspan=2)



if __name__ == '__main__':
    window = Tk()
    application = Product(window)
    window.mainloop()
